from datetime import datetime
from rich import print
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from openpyxl import Workbook, load_workbook
import os

Applications=[]

def Add(self):
    A_Id= int(input("Enter the Application ID :-"))
    Fname= input("Enter the First Name :-")
    Lname= input("Enter the Last Name :-")
    Cname= input("Enter the Company Name :-")
    CRole= input("Enter the Postion (Role) :-")
    Location= input("Enter the Company Location :-")
    Date = datetime.strptime(input("Enter the Date (eg : dd-mm-yyyy):- "), "%d-%m-%Y").date()
    Ctc = float(input("Enter the CTC (Package in LPA) :-"))
    Status= input("Enter the Status (Applied|Interview|Rejected|Selected):-")
    Applications.append([A_Id,Fname,Lname,Cname,CRole,Location,Date,Ctc,Status])
    print("[bold green]Applications Created Successfully !....[/bold green]")


def View(self):

    if len(Applications) == 0:
        print("[bold red]No Record Found[/bold red]")

    else:

        table = Table(title="Job Application Details",style="bright_cyan")

        table.add_column("Application ID", style="cyan")
        table.add_column("First Name", style="green")
        table.add_column("Last Name", style="green")
        table.add_column("Company", style="yellow")
        table.add_column("Position", style="blue")
        table.add_column("Location", style="magenta")
        table.add_column("Date", style="cyan")
        table.add_column("CTC", style="bright_green")
        table.add_column("Status", style="bright_blue")

        for i in Applications:
            if i[8] == "Applied":
                status = "[yellow]Applied[/yellow]"
            elif i[8] == "Interview":
                status = "[cyan]Interview[/cyan]"
            elif i[8] == "Rejected":
                status = "[red]Rejected[/red]"
            elif i[8] == "Selected":
                status = "[green]Selected[/green]"
            else:
                status = i[8]
            table.add_row(
                str(i[0]),
                i[1],
                i[2],
                i[3],
                i[4],
                i[5],
                str(i[6]),
                f"{i[7]} LPA",
                status
            )
        print(table)


# def View(self):
#     if (len(Applications))==0:
#         print("[bold red]No Record Found[/bold red]")
#     else:
#         print(f"[bold green]Application ID \tFirst Name\tLast Name\tCompany Name\tPosition\tLocation\tDate\tCTC\tStatus[/bold green ]")
#         for i in Applications:
#             print(f'{i[0]}\t{i[1]}\t{i[2]}\t{i[3]}\t{i[4]}\t{i[5]}\t{i[6]}\t{i[7]} LPA\t\t{i[8]}')


def Update(self):
    roll_a = int(console.input("[yellow]Enter the  Application ID to update :-[/yellow]"))
    flag = False
    for i in Applications:
        if i[0] == roll_a:
            print("[green]Found..[/green]")
            i[1]= input("Enter the First Name :-")
            i[2]= input("Enter the Last Name :-")
            i[3]= input("Enter the Company Name :-")
            i[4]= input("Enter the Postion (Role) :-")
            i[5]= input("Enter the Company Location :-")
            i[6]=datetime.strptime(input("Enter the Date (eg : dd-mm-yyyy):- "), "%d-%m-%Y").date()
            i[7]=float(input("Enter the CTC(Package) :-"))
            i[8]=input("Enter the Status (Applied|Interview|Rejected|Selected) :-")
            print("[bold green]Updated Application Successfully !.....[/bold green]")
            flag = True
            break
    if flag == False:
        print("[bold red]Not Found the Application ID [/bold red]")

def Search(self):
    Application_Id = int(console.input("[yellow]Enter the  Application ID to Search :-[/yellow]"))
    flag = False
    for i in Applications:
        if i[0] == Application_Id:
            print("[green]Found.....[/green]")
            print(f'{i[0]}\t{i[1]}\t{i[2]}\t\t{i[3]}\t{i[4]}\t{i[5]}\t{i[6]}\t{i[7]} LPA\t\t{i[8]}')
            flag = True
            break
    if flag == False:
        print("[bold red]Not Found the Application ID [/bold red]")

def Delete(self):
    NA_a = int(console.input("[yellow]Enter the  Application ID to delete :-[/yellow]"))
    flag = False
    for i in Applications:
        if i[0] == NA_a:
            Applications.remove(i)
            print("[bold green]Deleted Application Successfully !.....[/bold green]")
            flag = True
            break
    if flag == False:
        print("[bold red]Not Found the Application ID [/bold red]")


def ExportExcel(self):
    file_name = "JobApplications.xlsx"
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active
    else:

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Job Applications"

        # Add headings only for a new file
        headers = [
            "Application ID",
            "First Name",
            "Last Name",
            "Company",
            "Position",
            "Location",
            "Date",
            "CTC",
            "Status"
        ]
        sheet.append(headers)

    for i in Applications:

        sheet.append([
            i[0],
            i[1],
            i[2],
            i[3],
            i[4],
            i[5],
            str(i[6]),
            i[7],
            i[8]
        ])

    workbook.save(file_name)

    print("[bold green]Data exported successfully to Excel![/bold green]")


#-----------------------------
console = Console()
while True:
    console.print(
    Panel(
        "\n[bold cyan]Welcome to Job Application Tracker ![/bold cyan]\n",
        title="Python",
        border_style="green",
        width = 50 
    )
)
    choice=int(console.input("[dark_orange]1.Add Job Application\n2.View Application\n3.Update Application\n4.Search Application\n5.Delete Application\n6.Save the Data to Excel\n7.Exit[/dark_orange]\n[yellow]Enter Your Choice :-[/yellow] "))
    match choice:
        case 1:
            Add(self=Applications)
        case 2:
            View(self=Applications)
        case 3:
            Update(self=Applications)
        case 4:
            Search(self=Applications)
        case 5:
            Delete(self=Applications)
        case 6:
            ExportExcel(self=Applications)
        case 7:
            print("[bright_magenta ]Thank You......[/bright_magenta]")
            break
        case _:
            print("[bold red]Invalid Input[/bold red]")
            break

                    
                        
           
    
